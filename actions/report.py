# ============================================================
# actions/report.py
#
# Action: REPORT
#
# Compiles conversation decisions into a 6-sheet Excel workbook.
#
# FIX — "report not saving to outputs/":
#   The previous version only serialised to BytesIO (in-memory)
#   and never called wb.save(filepath). This version:
#     1. Resolves output_dir from session["output_dir"] or ./outputs
#     2. Calls wb.save(filepath) to write the file to disk
#     3. Also returns workbook_bytes so Streamlit can use download_button
#     4. Returns the absolute filepath in action_result so the formatter
#        can tell the user exactly where the file was saved.
#
# Uses openpyxl. No LLM calls. No state mutations. Pure Python.
# ============================================================

import io
import os
from datetime import datetime, timezone

HEADER_FILL_HEX = "1F4E79"
HEADER_FONT_HEX = "FFFFFF"
KEEP_FILL_HEX   = "E2EFDA"
DROP_FILL_HEX   = "FCE4D6"
FLAG_FILL_HEX   = "FFF2CC"
OVERRIDE_HEX    = "F4CCCC"


def run(state: dict) -> dict:
    try:
        import openpyxl
    except ImportError:
        return {"action_result": {"error": "openpyxl not installed. Run: pip install openpyxl"}}

    session      = state["session"]
    decisions    = state.get("decisions", {})
    decision_log = state.get("decision_log", [])

    feature_cols = session.get("feature_cols", [])
    verdict_df   = session.get("verdict_df")
    null_scan_df = session.get("null_scan_df")
    num_cols     = session.get("num_cols", [])
    cat_cols     = session.get("cat_cols", [])

    kept_cols        = sorted([c for c, d in decisions.items() if d == "keep"])
    dropped_cols     = sorted([c for c, d in decisions.items() if d == "drop"])
    pending_cols     = sorted([c for c in feature_cols if c not in decisions])
    override_entries = [e for e in decision_log if e.get("override") is True]
    null_indicator_cols = _find_null_indicators(null_scan_df, feature_cols)

    summary_counts = {
        "total_features" : len(feature_cols),
        "kept"           : len(kept_cols),
        "dropped"        : len(dropped_cols),
        "pending"        : len(pending_cols),
        "human_overrides": len(override_entries),
        "null_indicators": len(null_indicator_cols),
        "generated_at"   : _now(),
    }

    # ── Build workbook ────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheets_written = []

    def _add(title, fn, *args):
        ws = wb.create_sheet(title=title)
        fn(ws, *args, openpyxl=openpyxl)
        sheets_written.append(title)

    _add("Summary",         _write_summary,          summary_counts, null_indicator_cols)
    _add("All Decisions",   _write_all_decisions,    decision_log)
    _add("Human Overrides", _write_overrides,        override_entries, verdict_df)
    _add("Final Keep List", _write_keep_list,        kept_cols, verdict_df, num_cols, cat_cols)
    _add("Final Drop List", _write_drop_list,        dropped_cols, verdict_df, decision_log)
    _add("Null Indicators", _write_null_indicators,  null_indicator_cols, null_scan_df)

    # ── Resolve output directory ──────────────────────────────
    output_dir = session.get("output_dir") or os.path.join(os.getcwd(), "outputs")
    os.makedirs(output_dir, exist_ok=True)

    filename = f"ft_selection_report_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.abspath(os.path.join(output_dir, filename))

    # ── Save to disk ──────────────────────────────────────────
    wb.save(filepath)

    # ── Also return bytes for Streamlit ───────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return {
        "action_result": {
            "filepath"      : filepath,
            "filename"      : filename,
            "workbook_bytes": buf.getvalue(),
            "sheets_written": sheets_written,
            "summary_counts": summary_counts,
            "draft_mode"    : False,
        }
    }


# ── Sheet writers ─────────────────────────────────────────────

def _write_summary(ws, summary_counts, null_indicator_cols, openpyxl):
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    _header_row(ws, 1, ["Metric", "Value"], openpyxl)
    rows = [
        ("Total feature columns",  summary_counts["total_features"]),
        ("Kept",                   summary_counts["kept"]),
        ("Dropped",                summary_counts["dropped"]),
        ("Pending (no decision)",  summary_counts["pending"]),
        ("Human overrides",        summary_counts["human_overrides"]),
        ("Null indicator columns", summary_counts["null_indicators"]),
        ("Report generated at",    summary_counts["generated_at"]),
    ]
    for i, (metric, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=metric)
        ws.cell(row=i, column=2, value=value)
    start = len(rows) + 3
    ws.cell(row=start, column=1, value="Null Indicator Candidates").font = \
        openpyxl.styles.Font(bold=True)
    for j, col in enumerate(null_indicator_cols, start=start + 1):
        ws.cell(row=j, column=1, value=col)


def _write_all_decisions(ws, decision_log, openpyxl):
    headers = ["Column", "Decision", "Reason", "Source", "Timestamp", "Override"]
    for i, w in enumerate([22, 10, 42, 10, 26, 10], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    _header_row(ws, 1, headers, openpyxl)
    for i, entry in enumerate(decision_log, start=2):
        ws.cell(row=i, column=1, value=entry.get("col", ""))
        ws.cell(row=i, column=2, value=entry.get("decision", ""))
        ws.cell(row=i, column=3, value=entry.get("reason", ""))
        ws.cell(row=i, column=4, value=entry.get("source", ""))
        ws.cell(row=i, column=5, value=entry.get("timestamp", ""))
        ws.cell(row=i, column=6, value=str(entry.get("override", False)))
        if entry.get("override"):
            for c in range(1, 7):
                ws.cell(row=i, column=c).fill = openpyxl.styles.PatternFill(
                    fill_type="solid", fgColor=OVERRIDE_HEX)


def _write_overrides(ws, override_entries, verdict_df, openpyxl):
    headers = ["Column", "Bot Verdict", "User Decision", "Reason", "Timestamp"]
    for i, w in enumerate([22, 14, 14, 42, 26], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    _header_row(ws, 1, headers, openpyxl)
    verdict_map = _build_verdict_map(verdict_df)
    for i, entry in enumerate(override_entries, start=2):
        col = entry.get("col", "")
        ws.cell(row=i, column=1, value=col)
        ws.cell(row=i, column=2, value=verdict_map.get(col, "N/A"))
        ws.cell(row=i, column=3, value=entry.get("decision", ""))
        ws.cell(row=i, column=4, value=entry.get("reason", ""))
        ws.cell(row=i, column=5, value=entry.get("timestamp", ""))


def _write_keep_list(ws, kept_cols, verdict_df, num_cols, cat_cols, openpyxl):
    headers = ["Column", "Bot Verdict", "Confidence", "Type", "Risk Tag", "Profile"]
    for i, w in enumerate([22, 14, 12, 14, 18, 55], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    _header_row(ws, 1, headers, openpyxl)
    verdict_map    = _build_verdict_map(verdict_df)
    confidence_map = _build_field_map(verdict_df, ["confidence"])
    risk_tag_map   = _build_field_map(verdict_df, ["risk_tag"])
    profile_map    = _build_field_map(verdict_df, ["profile"])
    green_fill     = openpyxl.styles.PatternFill(fill_type="solid", fgColor=KEEP_FILL_HEX)
    for i, col in enumerate(kept_cols, start=2):
        col_type = "numeric" if col in num_cols else "categorical"
        conf     = confidence_map.get(col)
        ws.cell(row=i, column=1, value=col)
        ws.cell(row=i, column=2, value=verdict_map.get(col, "N/A"))
        ws.cell(row=i, column=3, value=round(float(conf), 1) if conf is not None else "N/A")
        ws.cell(row=i, column=4, value=col_type)
        ws.cell(row=i, column=5, value=risk_tag_map.get(col, ""))
        ws.cell(row=i, column=6, value=profile_map.get(col, ""))
        for c in range(1, 7):
            ws.cell(row=i, column=c).fill = green_fill


def _write_drop_list(ws, dropped_cols, verdict_df, decision_log, openpyxl):
    headers = ["Column", "Bot Verdict", "Drop Reason", "Source"]
    for i, w in enumerate([22, 14, 46, 12], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    _header_row(ws, 1, headers, openpyxl)
    verdict_map = _build_verdict_map(verdict_df)
    reason_map  = {}
    for entry in decision_log:
        if entry.get("decision") == "drop":
            reason_map[entry.get("col", "")] = entry
    red_fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor=DROP_FILL_HEX)
    for i, col in enumerate(dropped_cols, start=2):
        log_entry = reason_map.get(col, {})
        ws.cell(row=i, column=1, value=col)
        ws.cell(row=i, column=2, value=verdict_map.get(col, "N/A"))
        ws.cell(row=i, column=3, value=log_entry.get("reason", ""))
        ws.cell(row=i, column=4, value=log_entry.get("source", ""))
        for c in range(1, 5):
            ws.cell(row=i, column=c).fill = red_fill


def _write_null_indicators(ws, null_indicator_cols, null_scan_df, openpyxl):
    headers = ["Column", "Null Rate %", "Churn When Null %", "Churn When Present %", "Gap (pp)"]
    for i, w in enumerate([22, 12, 20, 22, 12], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    _header_row(ws, 1, headers, openpyxl)
    if not null_indicator_cols:
        ws.cell(row=2, column=1, value="No null indicator candidates found.")
        return
    yellow_fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor=FLAG_FILL_HEX)
    for i, col in enumerate(null_indicator_cols, start=2):
        r = _lookup_null_scan_row(col, null_scan_df)
        ws.cell(row=i, column=1, value=col)
        ws.cell(row=i, column=2, value=r.get("null_rate"))
        ws.cell(row=i, column=3, value=r.get("churn_null"))
        ws.cell(row=i, column=4, value=r.get("churn_present"))
        ws.cell(row=i, column=5, value=r.get("gap_pp"))
        for c in range(1, 6):
            ws.cell(row=i, column=c).fill = yellow_fill


# ── Helpers ───────────────────────────────────────────────────

def _header_row(ws, row, headers, openpyxl):
    fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor=HEADER_FILL_HEX)
    font = openpyxl.styles.Font(bold=True, color=HEADER_FONT_HEX)
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = fill
        cell.font = font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")


def _build_verdict_map(verdict_df) -> dict:
    if verdict_df is None:
        return {}
    try:
        if "column" in verdict_df.columns:
            return dict(zip(verdict_df["column"], verdict_df["verdict"].str.upper()))
        return verdict_df["verdict"].str.upper().to_dict()
    except Exception:
        return {}


def _build_field_map(verdict_df, keys) -> dict:
    if verdict_df is None:
        return {}
    for key in keys:
        try:
            if key in verdict_df.columns:
                if "column" in verdict_df.columns:
                    return dict(zip(verdict_df["column"], verdict_df[key]))
                return verdict_df[key].to_dict()
        except Exception:
            pass
    return {}


def _find_null_indicators(null_scan_df, feature_cols) -> list:
    """Uses 'col' field (run_null_signal_scan output schema)."""
    if null_scan_df is None or null_scan_df.empty:
        return []
    try:
        col_field = "col" if "col" in null_scan_df.columns else "column"
        gap_field = "gap_pp" if "gap_pp" in null_scan_df.columns else "gap"
        mask      = null_scan_df[gap_field].abs() > 3
        matched   = null_scan_df.loc[mask, col_field].tolist()
        return [c for c in matched if c in feature_cols]
    except Exception:
        return []


def _lookup_null_scan_row(col, null_scan_df) -> dict:
    """Handles 'col' field from run_null_signal_scan output."""
    if null_scan_df is None:
        return {}
    try:
        col_field = "col" if "col" in null_scan_df.columns else "column"
        rows = null_scan_df[null_scan_df[col_field] == col]
        if rows.empty:
            return {}
        r = rows.iloc[0]
        return {
            "null_rate"    : _sf(r, ["null_rate"]),
            "churn_null"   : _sf(r, ["churn_null", "churn_when_null"]),
            "churn_present": _sf(r, ["churn_present", "churn_when_present"]),
            "gap_pp"       : _sf(r, ["gap_pp", "gap"]),
        }
    except Exception:
        return {}


def _sf(row, keys) -> float | None:
    for k in keys:
        try:
            v = row[k]
            if v is not None and str(v) not in ("", "nan", "None"):
                return round(float(v), 2)
        except Exception:
            pass
    return None


def _now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")